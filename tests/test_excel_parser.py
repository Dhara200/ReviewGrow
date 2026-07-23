import io
import os
import tempfile
import unittest
import zipfile
from unittest.mock import patch

import pandas as pd
from flask import Flask, request
from openpyxl import Workbook

from app.config import Config
from app.utils import excel_parser
from app.utils.excel_parser import read_reviews_file


class ExcelParserTests(unittest.TestCase):
    def test_read_reviews_file_loads_xlsx_reviews(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "reviews.xlsx")
            pd.DataFrame(
                [
                    {
                        "source": "Google",
                        "rating": 5,
                        "review_text": "Excellent service",
                        "reviewer_name": "John",
                        "review_date": "20-06-2026",
                    }
                ]
            ).to_excel(upload_path, index=False)

            df = read_reviews_file(upload_path)

        self.assertEqual(len(df), 1)
        self.assertEqual(df.iloc[0]["review_text"], "Excellent service")
        self.assertEqual(df.iloc[0]["rating"], 5)

    def test_read_reviews_file_rejects_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "reviews.csv")
            pd.DataFrame([{"review_text": "Good"}]).to_csv(upload_path, index=False)

            with self.assertRaisesRegex(ValueError, "Excel file"):
                read_reviews_file(upload_path)

    def test_max_content_length_matches_upload_limit(self):
        self.assertGreater(Config.MAX_CONTENT_LENGTH, excel_parser.MAX_UPLOAD_BYTES)
        self.assertLessEqual(
            Config.MAX_CONTENT_LENGTH,
            excel_parser.MAX_UPLOAD_BYTES + 1024 * 1024,
        )

    def test_flask_rejects_request_above_max_content_length(self):
        app = Flask(__name__)
        app.config.from_object(Config)
        app.config["TESTING"] = True
        app.config["MAX_CONTENT_LENGTH"] = 1024

        @app.post("/upload-test")
        def upload_test():
            request.files.get("file")
            return "ok"

        response = app.test_client().post(
            "/upload-test",
            data={
                "file": (
                    io.BytesIO(b"x" * 2048),
                    "oversized.xlsx",
                )
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 413)
        response.close()

    def test_oversized_upload_is_rejected_before_parsing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "oversized.xlsx")
            with open(upload_path, "wb") as upload:
                upload.write(b"12345")

            with patch.object(excel_parser, "MAX_UPLOAD_BYTES", 4):
                with self.assertRaisesRegex(ValueError, "too large"):
                    read_reviews_file(upload_path)

    def test_corrupted_zip_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "corrupted.xlsx")
            with open(upload_path, "wb") as upload:
                upload.write(b"not a zip archive")

            with self.assertRaisesRegex(ValueError, "valid XLSX archive"):
                read_reviews_file(upload_path)

    def test_decompression_bomb_is_rejected_before_workbook_parsing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "bomb.xlsx")
            with zipfile.ZipFile(upload_path, "w", zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("[Content_Types].xml", "<Types />")
                archive.writestr("xl/workbook.xml", "A" * 100_000)

            with patch.object(excel_parser, "MAX_XLSX_COMPRESSION_RATIO", 5):
                with self.assertRaisesRegex(ValueError, "compression ratio"):
                    read_reviews_file(upload_path)

    def _write_workbook(self, path, *, sheets=1, rows=1, columns=1, value="ok"):
        workbook = Workbook()
        worksheet = workbook.active
        for column in range(1, columns + 1):
            worksheet.cell(row=1, column=column, value=value)
        for row in range(2, rows + 1):
            worksheet.cell(row=row, column=1, value=value)
        for index in range(1, sheets):
            workbook.create_sheet(f"Sheet{index + 1}")
        workbook.save(path)
        workbook.close()

    def test_excessive_rows_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "rows.xlsx")
            self._write_workbook(upload_path, rows=3)
            with patch.object(excel_parser, "MAX_ROWS", 2):
                with self.assertRaisesRegex(ValueError, "row limit"):
                    read_reviews_file(upload_path)

    def test_excessive_columns_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "columns.xlsx")
            self._write_workbook(upload_path, columns=3)
            with patch.object(excel_parser, "MAX_COLUMNS", 2):
                with self.assertRaisesRegex(ValueError, "column limit"):
                    read_reviews_file(upload_path)

    def test_excessive_worksheets_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "worksheets.xlsx")
            self._write_workbook(upload_path, sheets=3)
            with patch.object(excel_parser, "MAX_WORKSHEETS", 2):
                with self.assertRaisesRegex(ValueError, "worksheet limit"):
                    read_reviews_file(upload_path)

    def test_excessive_cell_text_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = os.path.join(temp_dir, "cell-text.xlsx")
            self._write_workbook(upload_path, value="too long")
            with patch.object(excel_parser, "MAX_CELL_TEXT_LENGTH", 4):
                with self.assertRaisesRegex(ValueError, "cell text"):
                    read_reviews_file(upload_path)


if __name__ == "__main__":
    unittest.main()
