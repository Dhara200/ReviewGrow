import os
import zipfile

import pandas as pd
from openpyxl import load_workbook

from app.utils.csv_parser import normalize_csv_row


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_WORKSHEETS = 10
MAX_ROWS = 10_000
MAX_COLUMNS = 50
MAX_CELL_TEXT_LENGTH = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
MAX_XLSX_ARCHIVE_MEMBERS = 1_000


def _validate_upload_size(upload_path):
    if os.path.getsize(upload_path) > MAX_UPLOAD_BYTES:
        raise ValueError("Excel file is too large. Maximum upload size is 10 MB.")


def _validate_xlsx_archive(upload_path):
    try:
        with zipfile.ZipFile(upload_path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                raise ValueError("Excel archive contains too many files.")

            names = [member.filename for member in members]
            if len(names) != len(set(names)):
                raise ValueError("Excel archive contains duplicate entries.")

            required_members = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required_members.issubset(names):
                raise ValueError("Excel file structure is invalid.")

            total_compressed = 0
            total_uncompressed = 0
            for member in members:
                normalized = member.filename.replace("\\", "/")
                parts = normalized.split("/")
                if normalized.startswith("/") or ".." in parts:
                    raise ValueError("Excel archive contains an invalid path.")
                if member.flag_bits & 0x1:
                    raise ValueError("Encrypted Excel archives are not supported.")

                total_compressed += member.compress_size
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("Excel archive expands beyond the safe limit.")

            if (
                total_uncompressed > 0
                and total_uncompressed
                > max(total_compressed, 1) * MAX_XLSX_COMPRESSION_RATIO
            ):
                raise ValueError("Excel archive compression ratio is unsafe.")

            if archive.testzip() is not None:
                raise ValueError("Excel archive is corrupted.")
    except (zipfile.BadZipFile, zipfile.LargeZipFile, EOFError) as error:
        raise ValueError("Excel file is not a valid XLSX archive.") from error


def _validate_workbook_limits(upload_path):
    try:
        workbook = load_workbook(
            upload_path,
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise ValueError("Excel workbook is corrupted or malformed.") from error

    try:
        if len(workbook.worksheets) > MAX_WORKSHEETS:
            raise ValueError(
                f"Excel workbook exceeds the {MAX_WORKSHEETS}-worksheet limit."
            )

        for worksheet in workbook.worksheets:
            if worksheet.max_row > MAX_ROWS:
                raise ValueError(
                    f"Excel worksheet exceeds the {MAX_ROWS}-row limit."
                )
            if worksheet.max_column > MAX_COLUMNS:
                raise ValueError(
                    f"Excel worksheet exceeds the {MAX_COLUMNS}-column limit."
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and len(cell.value) > MAX_CELL_TEXT_LENGTH
                    ):
                        raise ValueError(
                            "Excel cell text exceeds the safe length limit."
                        )
    finally:
        workbook.close()


def read_reviews_file(upload_path):
    extension = os.path.splitext(upload_path)[1].lower()

    if extension not in {".xlsx", ".xls"}:
        raise ValueError("Unsupported file type. Please upload an Excel file.")

    _validate_upload_size(upload_path)

    if extension == ".xlsx":
        _validate_xlsx_archive(upload_path)
        _validate_workbook_limits(upload_path)

    try:
        excel_file = pd.ExcelFile(upload_path)
        if len(excel_file.sheet_names) > MAX_WORKSHEETS:
            raise ValueError(
                f"Excel workbook exceeds the {MAX_WORKSHEETS}-worksheet limit."
            )
        dataframe = pd.read_excel(excel_file)
    except ValueError:
        raise
    except Exception as error:
        raise ValueError("Excel workbook is corrupted or malformed.") from error
    finally:
        if "excel_file" in locals():
            excel_file.close()

    if len(dataframe.index) > MAX_ROWS:
        raise ValueError(f"Excel worksheet exceeds the {MAX_ROWS}-row limit.")
    if len(dataframe.columns) > MAX_COLUMNS:
        raise ValueError(f"Excel worksheet exceeds the {MAX_COLUMNS}-column limit.")

    for value in dataframe.select_dtypes(include=["object"]).to_numpy().flat:
        if isinstance(value, str) and len(value) > MAX_CELL_TEXT_LENGTH:
            raise ValueError("Excel cell text exceeds the safe length limit.")

    return dataframe


def iter_normalized_rows(df):
    for _, row in df.iterrows():
        yield normalize_csv_row(row)
